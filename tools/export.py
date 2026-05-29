"""Excel(.xlsx) 출력 유틸리티.

SearchResult를 받아 Summary 시트 + Structures 시트로 구성된 워크북을 저장한다.
GPCR 타깃이면 확장 컬럼(State / Ligand / Modality 등)과 조건부 색상을 적용한다.
"""

from __future__ import annotations

import datetime
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from models.schemas import Citation, PDBEntry, SearchResult
from tools.pdb import format_method

def _default_output_dir() -> Path:
    """Excel 저장 디렉토리를 결정한다 (항상 절대 경로).

    MCP 서버는 Claude Desktop이 임의의 작업 디렉토리(보통 `/`)에서 실행하므로,
    상대 경로를 쓰면 권한 없는 위치에 저장을 시도해 실패한다. 따라서 항상
    절대 경로를 반환한다.

    - 환경변수 PDB_MCP_OUTPUT_DIR 가 있으면 그 경로
      (.mcpb 배포에서는 Claude Desktop이 user_config 값을 이 변수로 전달한다)
    - 없으면 사용자가 쉽게 찾을 수 있는 ~/Documents/PDBMCP
    """
    env_dir = os.environ.get("PDB_MCP_OUTPUT_DIR")
    if env_dir:
        return Path(env_dir).expanduser()
    return Path.home() / "Documents" / "PDBMCP"

# 공통 서식 상수
_HEADER_FILL = PatternFill("solid", fgColor="1E293B")  # 진한 남색
_EVEN_FILL = PatternFill("solid", fgColor="F8FAFC")  # 연한 회색
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_LINK_FONT = Font(color="2B6CB0", underline="single")
_LABEL_FONT = Font(bold=True)

# State 조건부 색상
_STATE_FILLS = {
    "Active": PatternFill("solid", fgColor="DCFCE7"),       # 연한 초록
    "Inactive": PatternFill("solid", fgColor="FEF2F2"),     # 연한 빨강
    "Intermediate": PatternFill("solid", fgColor="FEF9C3"), # 연한 노랑
}

# Ligand modality 조건부 색상
_MODALITY_FILLS = {
    "Agonist": PatternFill("solid", fgColor="DCFCE7"),
    "Partial agonist": PatternFill("solid", fgColor="DCFCE7"),
    "Antagonist": PatternFill("solid", fgColor="FEF2F2"),
    "Inverse agonist": PatternFill("solid", fgColor="FFF7ED"),
}

# 컬럼 정의 (순서 고정)
_BASIC_COLUMNS = [
    "PDB ID", "Resolution (Å)", "Method", "Released Date", "Entry Title",
    "Paper Title", "Authors", "Journal", "Year", "Citation (ACS)", "DOI", "PMID",
]
_GPCR_COLUMNS = [
    "Method", "PDB ID", "Res. (Å)", "Pref. chain", "State", "Ligand",
    "Ligand modality", "Sign. prot.", "Fusion protein", "Antibody",
    "Year", "Citation (ACS)", "DOI", "PMID",
]

# 컬럼 너비 상한 / Citation 컬럼 전용 너비
_MAX_WIDTH = 60
_CITATION_WIDTH = 75


def _year_of(entry: PDBEntry) -> int | None:
    """citation.year 우선, 없으면 released_date 앞 4자리에서 연도를 추출."""
    if entry.citation and entry.citation.year:
        return entry.citation.year
    if entry.released_date and entry.released_date[:4].isdigit():
        return int(entry.released_date[:4])
    return None


def format_acs_citation(citation: Citation | None) -> str:
    """Citation 객체를 ACS 스타일 인용문 문자열로 변환한다.

    형식:
      Author1, F. M.; Author2, F. M. Title. J. Abbrev. Year, Vol, PageFirst–PageLast. DOI: xx.

    필드가 없는 부분은 생략하고 나머지로 조합한다.
    """
    if citation is None:
        return ""

    parts: list[str] = []

    # 저자 (이미 세미콜론 구분, 마침표로 끝맺음)
    if citation.authors:
        parts.append(citation.authors.rstrip(".") + ".")

    # 논문 제목
    if citation.title:
        parts.append(citation.title.rstrip(".") + ".")

    # 저널명 Year, Vol, Pages.
    journal_part = ""
    if citation.journal:
        journal_part += citation.journal
    if citation.year:
        journal_part += f" {citation.year}"
    if citation.volume:
        journal_part += f", {citation.volume}"
    if citation.page_first:
        if citation.page_last and citation.page_last != citation.page_first:
            journal_part += f", {citation.page_first}–{citation.page_last}"
        else:
            journal_part += f", {citation.page_first}"
    if journal_part:
        parts.append(journal_part.strip() + ".")

    # DOI
    if citation.doi:
        parts.append(f"DOI: {citation.doi}.")

    return " ".join(parts)


def export_to_excel(
    result: SearchResult, output_dir: Path | str | None = None
) -> str:
    """SearchResult를 .xlsx 파일로 저장하고 절대 경로를 반환한다.

    output_dir 미지정 시 _default_output_dir()(환경변수 PDB_MCP_OUTPUT_DIR 또는
    ~/Documents/PDBMCP)에 저장한다.
    파일명 규칙: {GENE_NAME}_{ACCESSION}_structures_{YYYYMMDD}.xlsx
    """
    output_dir = Path(output_dir).expanduser() if output_dir else _default_output_dir()
    output_dir.mkdir(parents=True, exist_ok=True)

    gene = (
        result.uniprot.gene_name
        or result.uniprot.entry_name
        or "UNKNOWN"
    ).replace(" ", "_")
    date_str = datetime.date.today().strftime("%Y%m%d")
    filename = f"{gene}_{result.uniprot.accession}_structures_{date_str}.xlsx"
    filepath = output_dir / filename

    wb = Workbook()
    _build_summary_sheet(wb.active, result)
    data_sheet = wb.create_sheet("Structures")
    if result.uniprot.is_gpcr:
        _build_gpcr_sheet(data_sheet, result)
    else:
        _build_basic_sheet(data_sheet, result)
    wb.save(filepath)

    return str(filepath.resolve())


def export_family_to_excel(
    family_name: str,
    results: list[SearchResult],
    output_dir: Path | str | None = None,
) -> str:
    """여러 SearchResult를 Summary + 타깃별 시트로 저장한다.

    패밀리 단위 요청에서 Claude가 화면 표를 다시 xlsx로 만들지 않도록, 서버가
    단일 타깃 Excel과 동일한 Resolution 서식/ACS Citation 컬럼을 유지해 저장한다.
    """
    if not results:
        raise ValueError("저장할 검색 결과가 없습니다.")

    output_dir = Path(output_dir).expanduser() if output_dir else _default_output_dir()
    output_dir.mkdir(parents=True, exist_ok=True)

    label = _filename_token(family_name or "_".join(r.uniprot.gene_name or r.query for r in results))
    date_str = datetime.date.today().strftime("%Y%m%d")
    filepath = output_dir / f"{label}_family_structures_{date_str}.xlsx"

    wb = Workbook()
    _build_family_summary_sheet(wb.active, family_name, results)

    used_names = {"Summary"}
    for result in results:
        sheet_name = _unique_sheet_name(
            _sheet_name(result.uniprot.gene_name or result.query or result.uniprot.accession),
            used_names,
        )
        used_names.add(sheet_name)
        ws = wb.create_sheet(sheet_name)
        if result.uniprot.is_gpcr:
            _build_gpcr_sheet(ws, result)
        else:
            _build_basic_sheet(ws, result)

    wb.save(filepath)
    return str(filepath.resolve())


def _filename_token(value: str) -> str:
    """파일명에 안전한 토큰으로 변환한다."""
    token = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in value)
    token = "_".join(part for part in token.split("_") if part)
    return token[:80] or "PDB"


def _sheet_name(value: str) -> str:
    """Excel 시트명 제약에 맞게 변환한다."""
    cleaned = "".join(ch if ch not in r'[]:*?/\\' else "_" for ch in value).strip()
    return (cleaned or "Target")[:31]


def _unique_sheet_name(base: str, used: set[str]) -> str:
    """중복되지 않는 Excel 시트명을 만든다."""
    if base not in used:
        return base
    for i in range(2, 100):
        suffix = f"_{i}"
        candidate = base[: 31 - len(suffix)] + suffix
        if candidate not in used:
            return candidate
    raise ValueError("시트명이 너무 많이 중복됩니다.")


def _build_family_summary_sheet(
    ws: Worksheet,
    family_name: str,
    results: list[SearchResult],
) -> None:
    """패밀리 workbook의 Summary 시트를 구성한다."""
    ws.title = "Summary"
    columns = [
        "Target", "UniProt", "Type", "Total PDB", "Included",
        "GPCRdb matched", "Best resolution", "Latest structure",
    ]
    widths = _header_row(ws, columns)

    for idx, result in enumerate(results, start=2):
        u = result.uniprot
        even = idx % 2 == 0
        _put(ws, idx, 1, u.gene_name or result.query, widths, even)
        _put(ws, idx, 2, u.accession, widths, even)
        _put(ws, idx, 3, "GPCR" if u.is_gpcr else "Basic", widths, even)
        _put(ws, idx, 4, result.total_count, widths, even, center=True)
        _put(ws, idx, 5, len(result.structures), widths, even, center=True)
        _put(ws, idx, 6, result.gpcrdb_count or 0 if u.is_gpcr else "-", widths, even, center=True)
        _put(ws, idx, 7, _best_resolution_summary(result.structures), widths, even)
        _put(ws, idx, 8, _latest_structure_summary(result.structures), widths, even)

    meta_row = len(results) + 4
    ws.cell(row=meta_row, column=1, value="조회 일시").font = _LABEL_FONT
    ws.cell(row=meta_row, column=2, value=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    _finalize(ws, columns, widths, len(results))


def _best_resolution_summary(structures: list[PDBEntry]) -> str:
    """최고 해상도 요약 문자열."""
    with_res = [entry for entry in structures if entry.resolution is not None]
    if not with_res:
        return "-"
    # with_res가 이미 None을 거른 상태이므로 `or 0.0` fallback은 불필요.
    best = min(with_res, key=lambda entry: entry.resolution)
    return f"{best.resolution:.2f} Å ({best.pdb_id})"


def _latest_structure_summary(structures: list[PDBEntry]) -> str:
    """최신 구조 요약 문자열."""
    with_date = [entry for entry in structures if entry.released_date]
    if not with_date:
        return "-"
    latest = max(with_date, key=lambda entry: entry.released_date or "")
    return f"{latest.pdb_id} ({latest.released_date})"


def _build_summary_sheet(ws: Worksheet, result: SearchResult) -> None:
    """요약 정보 시트를 구성한다."""
    ws.title = "Summary"
    u = result.uniprot

    title_cell = ws.cell(
        row=1, column=1, value=f"{u.gene_name or u.entry_name} 구조 검색 요약"
    )
    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")

    rows: list[tuple[str, object]] = [
        ("타겟 검색어", result.query),
        ("UniProt Accession", u.accession),
        ("Entry Name", u.entry_name),
        ("단백질명", u.protein_name),
        ("유전자명", u.gene_name or "-"),
        ("Organism", u.organism or "-"),
        ("GPCR 여부", "예 (GPCRdb 연동)" if u.is_gpcr else "아니오"),
        ("총 PDB 구조 수", result.total_count),
        ("수록 구조 수", len(result.structures)),
    ]
    if u.is_gpcr:
        rows.append(("GPCRdb 기준 구조 수", result.gpcrdb_count or 0))
    rows.append(("조회 일시", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

    for i, (label, value) in enumerate(rows, start=3):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font = _LABEL_FONT
        label_cell.fill = _EVEN_FILL
        ws.cell(row=i, column=2, value=value)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60


def _header_row(ws: Worksheet, columns: list[str]) -> list[int]:
    """헤더 행을 기록하고 컬럼 너비 추적 리스트를 반환한다."""
    for col, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    return [len(name) + 2 for name in columns]


def _put(
    ws: Worksheet,
    row: int,
    col: int,
    value,
    widths: list[int],
    even: bool,
    *,
    link: str | None = None,
    fill: PatternFill | None = None,
    num_format: str | None = None,
    center: bool = False,
    wrap: bool = False,
) -> None:
    """셀 하나를 기록한다. fill이 주어지면 짝수행 줄무늬보다 우선 적용."""
    cell = ws.cell(row=row, column=col, value=value if value not in ("", None) else None)
    applied_fill = fill if fill is not None else (_EVEN_FILL if even else None)
    if applied_fill is not None:
        cell.fill = applied_fill
    if link and value not in ("", None):
        cell.hyperlink = link
        cell.font = _LINK_FONT
    if num_format:
        cell.number_format = num_format
    if wrap:
        cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
    elif center:
        cell.alignment = Alignment(horizontal="center")

    text_len = len(str(value)) if value not in ("", None) else 0
    if text_len + 2 > widths[col - 1]:
        widths[col - 1] = text_len + 2


def _finalize(
    ws: Worksheet, columns: list[str], widths: list[int], n_rows: int
) -> None:
    """컬럼 너비 자동 조정 + 틀 고정 + 필터를 적용한다.

    Citation (ACS) 컬럼은 긴 인용문을 담으므로 전용 너비(_CITATION_WIDTH)를 쓴다.
    """
    for i, width in enumerate(widths, start=1):
        if columns[i - 1] == "Citation (ACS)":
            ws.column_dimensions[get_column_letter(i)].width = _CITATION_WIDTH
        else:
            ws.column_dimensions[get_column_letter(i)].width = min(width, _MAX_WIDTH)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(widths))}{max(n_rows + 1, 1)}"


def _build_basic_sheet(ws: Worksheet, result: SearchResult) -> None:
    """비GPCR 타깃용 기본 컬럼 시트."""
    widths = _header_row(ws, _BASIC_COLUMNS)

    for idx, entry in enumerate(result.structures):
        row = idx + 2
        even = idx % 2 == 1
        cit = entry.citation

        _put(ws, row, 1, entry.pdb_id, widths, even,
             link=f"https://www.rcsb.org/structure/{entry.pdb_id}")
        if entry.resolution is None:
            _put(ws, row, 2, "N/A", widths, even, center=True)
        else:
            # 항상 float + '0.00' 서식으로 소수점 2자리 통일
            _put(ws, row, 2, round(float(entry.resolution), 2), widths, even,
                 num_format="0.00", center=True)
        _put(ws, row, 3, format_method(entry.method), widths, even)
        _put(ws, row, 4, entry.released_date or "", widths, even)
        _put(ws, row, 5, entry.title or "", widths, even)
        _put(ws, row, 6, (cit.title if cit else "") or "", widths, even)
        _put(ws, row, 7, (cit.authors if cit else "") or "", widths, even)
        _put(ws, row, 8, (cit.journal if cit else "") or "", widths, even)
        _put(ws, row, 9, _year_of(entry) or "", widths, even)
        _put(ws, row, 10, format_acs_citation(cit), widths, even, wrap=True)
        doi = cit.doi if cit else None
        _put(ws, row, 11, doi or "", widths, even,
             link=f"https://doi.org/{doi}" if doi else None)
        pmid = cit.pmid if cit else None
        _put(ws, row, 12, pmid or "", widths, even,
             link=f"https://pubmed.ncbi.nlm.nih.gov/{pmid}" if pmid else None)

    _finalize(ws, _BASIC_COLUMNS, widths, len(result.structures))


def _build_gpcr_sheet(ws: Worksheet, result: SearchResult) -> None:
    """GPCR 타깃용 확장 컬럼 시트 (State / Modality 조건부 색상 포함)."""
    widths = _header_row(ws, _GPCR_COLUMNS)

    for idx, entry in enumerate(result.structures):
        row = idx + 2
        even = idx % 2 == 1
        cit = entry.citation

        _put(ws, row, 1, format_method(entry.method), widths, even)
        _put(ws, row, 2, entry.pdb_id, widths, even,
             link=f"https://www.rcsb.org/structure/{entry.pdb_id}")
        if entry.resolution is None:
            _put(ws, row, 3, "N/A", widths, even, center=True)
        else:
            # 항상 float + '0.00' 서식으로 소수점 2자리 통일
            _put(ws, row, 3, round(float(entry.resolution), 2), widths, even,
                 num_format="0.00", center=True)
        _put(ws, row, 4, entry.pref_chain or "-", widths, even, center=True)

        # State — 조건부 색상
        state_fill = _STATE_FILLS.get(entry.state or "")
        _put(ws, row, 5, entry.state or "-", widths, even,
             fill=state_fill, center=True)

        _put(ws, row, 6, entry.ligand or "-", widths, even)

        # Ligand modality — 조건부 색상
        modality_fill = _MODALITY_FILLS.get(entry.ligand_modality or "")
        _put(ws, row, 7, entry.ligand_modality or "-", widths, even,
             fill=modality_fill)

        _put(ws, row, 8, entry.signaling_protein or "-", widths, even, center=True)
        _put(ws, row, 9, entry.fusion_protein or "-", widths, even, center=True)
        _put(ws, row, 10, entry.antibody or "-", widths, even, center=True)
        _put(ws, row, 11, _year_of(entry) or "", widths, even, center=True)
        _put(ws, row, 12, format_acs_citation(cit), widths, even, wrap=True)

        doi = cit.doi if cit else None
        _put(ws, row, 13, doi or "", widths, even,
             link=f"https://doi.org/{doi}" if doi else None)
        pmid = cit.pmid if cit else None
        _put(ws, row, 14, pmid or "", widths, even,
             link=f"https://pubmed.ncbi.nlm.nih.gov/{pmid}" if pmid else None)

    _finalize(ws, _GPCR_COLUMNS, widths, len(result.structures))
