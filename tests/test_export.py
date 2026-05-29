"""tools/export.py 테스트 — ACS 인용 포맷 + Excel 출력 (오프라인 전용)."""

import tempfile

from openpyxl import load_workbook

from models.schemas import Citation, PDBEntry, SearchResult, UniProtResult
from tools.export import export_family_to_excel, export_to_excel, format_acs_citation


# --------------------------------------------------------------------------
# format_acs_citation
# --------------------------------------------------------------------------

def test_format_acs_citation_full():
    cit = Citation(
        title="Structure of the receptor",
        authors="Huang, J.; Chen, S.; Zhang, J. J. et al.",
        journal="Nature",
        year=2020,
        volume="579",
        page_first="303",
        page_last="308",
        doi="10.1038/s41586-020-1968-7",
    )
    assert format_acs_citation(cit) == (
        "Huang, J.; Chen, S.; Zhang, J. J. et al. "
        "Structure of the receptor. Nature 2020, 579, 303–308. "
        "DOI: 10.1038/s41586-020-1968-7."
    )


def test_format_acs_citation_partial():
    """필드가 없으면 해당 부분을 생략한다."""
    assert format_acs_citation(Citation(journal="Science", year=2022)) == "Science 2022."


def test_format_acs_citation_single_page():
    """page_last가 없으면 시작 페이지만 표기."""
    cit = Citation(title="T", journal="J", year=2021, volume="5", page_first="100")
    assert format_acs_citation(cit) == "T. J 2021, 5, 100."


def test_format_acs_citation_none():
    assert format_acs_citation(None) == ""


# --------------------------------------------------------------------------
# Excel 출력 — Citation (ACS) 컬럼
# --------------------------------------------------------------------------

def test_export_includes_acs_column():
    result = SearchResult(
        query="EGFR",
        uniprot=UniProtResult(
            accession="P00533", entry_name="EGFR_HUMAN",
            protein_name="Epidermal growth factor receptor", gene_name="EGFR",
        ),
        structures=[
            PDBEntry(
                pdb_id="7T9K", resolution=2.45, method="EM",
                released_date="2021-12-29",
                citation=Citation(
                    title="A paper", authors="Lee, K.; Kim, J. et al.",
                    journal="Science", year=2022, volume="375",
                    page_first="760", page_last="764", doi="10.1126/science.x",
                ),
            )
        ],
        total_count=1,
    )
    with tempfile.TemporaryDirectory() as tmp:
        path = export_to_excel(result, output_dir=tmp)
        ws = load_workbook(path)["Structures"]

        headers = [c.value for c in ws[1]]
        assert "Citation (ACS)" in headers

        acs_col = headers.index("Citation (ACS)") + 1
        cell = ws.cell(row=2, column=acs_col)
        assert cell.value is not None
        assert "Science 2022, 375, 760–764" in cell.value
        assert "DOI: 10.1126/science.x" in cell.value
        assert cell.alignment.wrap_text is True


def test_export_family_uses_server_workbook_format():
    result = SearchResult(
        query="HTR2A",
        uniprot=UniProtResult(
            accession="P28223",
            entry_name="5HT2A_HUMAN",
            protein_name="5-hydroxytryptamine receptor 2A",
            gene_name="HTR2A",
            is_gpcr=True,
        ),
        structures=[
            PDBEntry(
                pdb_id="6A93",
                resolution=2.9,
                method="X-RAY DIFFRACTION",
                released_date="2018-07-18",
                state="Inactive",
                ligand="Risperidone",
                ligand_modality="Antagonist",
                is_gpcr=True,
                citation=Citation(
                    title="Structural insights",
                    authors="Kim, J.; Lee, S. et al.",
                    journal="Cell",
                    year=2018,
                    volume="172",
                    page_first="719",
                    page_last="730",
                    doi="10.1016/j.cell.x",
                ),
            )
        ],
        total_count=1,
        gpcrdb_count=1,
    )
    with tempfile.TemporaryDirectory() as tmp:
        path = export_family_to_excel("HTR2", [result], output_dir=tmp)
        wb = load_workbook(path)
        assert wb.sheetnames == ["Summary", "HTR2A"]

        ws = wb["HTR2A"]
        headers = [c.value for c in ws[1]]
        assert "Citation (ACS)" in headers
        assert ws.cell(row=2, column=3).number_format == "0.00"

        acs_col = headers.index("Citation (ACS)") + 1
        assert "Cell 2018, 172, 719–730" in ws.cell(row=2, column=acs_col).value
